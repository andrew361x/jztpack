# -*-    coding:utf-8       -*-
import os
import os.path
import shutil
from collections import OrderedDict
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os
from .utils import convertXls2Xlsx, jztProReportPicture
from operator import itemgetter

from collections import OrderedDict
from datetime import datetime

# for   readJztSimpleBacktestReportTradeDetail
from PyQt5.QtCore import QFile, QIODevice, QTextStream


class jztProBacktestTradedetailTable:
    """封装金字塔专业回测报告交易明细表格的接口"""

    def __init__(self, fileName=None, colWidthDict=None):
        """colWidthDict各列数据的列宽,品种列可能需要重新设置"""
        self._fileName = fileName
        # self._widthDict = OrderedDict({"时间":23,"名称":15,"类型":12,"交易价":17,"交易量":10,"收益":14,"幅度":23,"资产":16,"最大回撤":10})#如果所有字符都是英文就是这个配置
        self._widthDict = OrderedDict({
            "序号": 8,
            "品种": 11,
            "交易类型": 8,
            "时间": 23,
            "数量": 8,
            "交易效率": 12,
            "开仓价格": 12,
            "平仓价格": 17,
            "盈亏": 10,
            "盈亏金额": 14,
            "最大回撤": 20,
            "累计盈亏金额": 16,
            "手续费": 10,
            "滑价成本": 12,
            "最佳平仓价": 14,
            "最差平仓价": 14,
            "最大浮盈": 14,
            "最大浮亏": 14,
            "开仓效率": 12,
            "平仓效率": 12,
            "有效盈亏": 13,
            "资产": 16,
            "持仓周期": 12,
            "周期均盈利": 14,
            "保证金占用": 11
        })  # 因为配置中有中文，但是并没有按照两个计算，把每个字符当一个英文字符统计

        if not colWidthDict is None:
            self._widthDict.update(colWidthDict)

    def showConfig(self):
        print(self._widthDict)

    def updateConfig(self, config):
        self._widthDict.update(config)

    def read(self):
        """ 读取金字塔专业回测报告交易明细表格
        默认表头:
        序号    品种           交易类型    时间                   数量    交易效率    开仓价格    平仓价格/均价    盈亏%     盈亏金额      最大回撤(单品种)    
        累计盈亏金额    手续费    滑价成本    最佳平仓价    最差平仓价    最大浮盈      最大浮亏      开仓效率    平仓效率    有效盈亏%    资产(单品种)    持仓周期    
        周期均盈利    保证金占用
        默认宽度:"序号":8,"品种":15,"交易类型":5,"时间":23,"数量":8,"交易效率":12,"开仓价格":12,"平仓价格":17,"盈亏":10,
        "盈亏金额":14,"最大回撤":20,"累计盈亏金额":16,"手续费":10,"滑价成本":12,"最佳平仓价":14,"最差平仓价":14,"最大浮盈":14,     
         "最大浮亏":14,"开仓效率":12,"平仓效率":12,"有效盈亏":13,"资产":16,"持仓周期":12,   
        "周期均盈利":14,"保证金占用":11
        """
        _names = list(self._widthDict.keys())
        _widths = list(self._widthDict.values())
        fh = open(self._fileName)
        data = pd.read_fwf(fh,
                           widths=_widths,
                           names=_names,
                           encoding="GB2312",
                           skiprows=(0, ))
        #print(data.iloc[0])

        data["序号"] = [int(i) for i in data["序号"]]
        data["品种"] = [str(i.strip()) for i in data["品种"]]
        data["交易类型"] = [str(i) for i in data["交易类型"]]
        DTFORMAT = r"%Y/%m/%d %H:%M:%S"
        data["时间"] = [datetime.strptime(i, DTFORMAT) for i in data["时间"]]
        data["数量"] = [int(i) for i in data["数量"]]
        data["交易效率"] = [
            float(str(i).replace("%", "")) / 100. for i in data["交易效率"]
        ]
        data["开仓价格"] = [float(str(i).strip()) for i in data["开仓价格"]]
        data["平仓价格"] = [float(str(i).split("/")[0]) for i in data["平仓价格"]]
        data["盈亏"] = [
            float(str(i).replace("%", "")) / 100. for i in data["盈亏"]
        ]
        data["盈亏金额"] = [float(str(i).replace(",", "")) for i in data["盈亏金额"]]
        data["最大回撤"] = [
            float(str(i).replace("%", "")) / 100. for i in data["最大回撤"]
        ]
        data["累计盈亏金额"] = [
            float(str(i).replace(",", "")) for i in data["累计盈亏金额"]
        ]
        data["手续费"] = [float(str(i).strip()) for i in data["手续费"]]
        data["滑价成本"] = [float(str(i).strip()) for i in data["滑价成本"]]
        data["最佳平仓价"] = [float(str(i).strip()) for i in data["最佳平仓价"]]
        data["最差平仓价"] = [float(str(i).strip()) for i in data["最差平仓价"]]
        data["最大浮盈"] = [float(str(i).replace(",", "")) for i in data["最大浮盈"]]
        data["最大浮亏"] = [float(str(i).replace(",", "")) for i in data["最大浮亏"]]
        data["开仓效率"] = [
            float(str(i).replace("%", "")) / 100. for i in data["开仓效率"]
        ]
        data["平仓效率"] = [
            float(str(i).replace("%", "")) / 100. for i in data["平仓效率"]
        ]
        data["有效盈亏"] = [
            float(str(i).replace("%", "")) / 100. for i in data["有效盈亏"]
        ]
        data["资产"] = [float(str(i).replace(",", "")) for i in data["资产"]]
        data["持仓周期"] = [float(i) for i in data["持仓周期"]]
        data["周期均盈利"] = [float(str(i).replace(",", "")) for i in data["周期均盈利"]]
        data["保证金占用"] = [float(str(i).replace(",", "")) for i in data["保证金占用"]]
        return data


class jztProBactTestXlsxFile:
    """针对金字塔单策略专业回测报告导出的xls文件封装的模块"""

    def __init__(self, fullpath=None):

        self.dirname = os.path.dirname(fullpath)
        self.filename = os.path.basename(fullpath)
        self.labelImgdict = {
            "资金曲线": "image1",
            "资金及水下回撤曲线": "image2",
            "资金升水及回撤曲线": "image3",
            "资金升水及回撤幅度曲线": "image4",
            "平仓收益曲线": "image5",
            "平仓收益及回撤": "image6",
            "多头平仓收益及回撤": "image7",
            "空头平仓收益及回撤": "image8",
            "平仓盈亏散点图": "image9",
            "有效盈亏散点图": "image10",
            "平仓盈亏分布图": "image11",
            "最大浮盈散点图": "image12",
            "最大浮盈-平仓盈亏柱状图": "image13",
            "最大浮盈-平仓盈亏散点图": "image14",
            "最大浮亏散点图": "image15",
            "最大浮亏-平仓盈亏柱状图": "image16",
            "最大浮亏-平仓盈亏散点图": "image17",
            "月净值收益与回撤柱状图": "image18",
            "月净值收益与回撤幅度柱状图": "image19",
            "月平仓收益与回撤柱状图": "image20",
            "月平仓收益与回撤幅度柱状图": "image21"
        }
        self.imgLabeldict = {v: k for k, v in self.labelImgdict.items()}

        if self.filename.lower().endswith("xls"):
            convertXls2Xlsx(fullpath)
            self.filename = self.filename.lower() + "x"

        self.fullpath = os.path.join(self.dirname, self.filename)
        self.sheetNames = None
        self._open()

        self.summaryDict = {}
        self.dataList = set(self.sheetNames).difference(
            set(self.labelImgdict.keys()))
        self.funcList = {
            '总体概要': self.readGeneralOutline,
            '收益风险分析': self.readRiskReturnAnalysis,
            '时间仓位分析': self.readTimeHoldingAnalysis,
            '交易明细': self.readTradeDetails,
            '交易总体分析': self.readTradeOverallAnalysis,
            '连续盈亏分析': self.readContinuousPLAnalysis,
            '离群交易': self.readOutlierTrading,
            '最大浮赢浮亏': self.readMaxFloatingPL,
            '日交易分析': self.readDailyTradeAnalysis,
            '月绩效分析': self.readMonthlyPerformanceAnalysis,
            '月度分析': self.readMonthlyAnalysis,
            '年度分析': self.readYearlyAnalysis
        }
        self.picList = list(self.labelImgdict.keys())

    def _open(self):
        """使用openpyxl打开excel文件"""
        # 默认可读写，只读操作
        self.wb = load_workbook(self.fullpath, read_only=True)
        self.sheetNames = self.wb.sheetnames

    def closeExcel(self):
        if self.wb:
            self.wb.close()
            self.wb = None
        # 操作结束 无法彻底解绑excel的问题还是没有解决掉

    def showSheetNames(self):
        """获取所有sheetname"""
        return self.sheetNames

    def getReport(self, labelList="all"):
        """获取所有数据 #labellist 是要计算的sheet名字组成的list """
        if labelList == "all":
            [v(sheetName=k) for k, v in self.funcList.items()]
        else:
            [v(sheetName=k) for k, v in self.funcList if k in labelList]

    def __len__(self):
        return len(self.summaryDict)

    def __getitem__(self, sheetName=None):
        return self.summaryDict[sheetName]

    def readGeneralOutline(self, sheetName=None):
        """读取金字塔专业回测报告中的总体概要"""
        sheet = self.wb[sheetName]
        generalOutline = {}
        # 读取第2行到31行的数据
        lines = range(2, 32)
        for i in lines:
            if sheet["A" + str(i)].value:
                generalOutline[sheet["A" + str(i)].value] = {
                    "所有交易": sheet["B" + str(i)].value,
                    "多头交易": sheet["C" + str(i)].value,
                    "空头交易": sheet["D" + str(i)].value
                }

        # 读取第34行到40行的数据
        lines = range(34, 41)
        for i in lines:
            line = sheet["A" + str(i)].value.split("：")
            generalOutline[line[0]] = line[1]
        self.summaryDict.update({sheetName: generalOutline})
        return generalOutline

    def readRiskReturnAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的收益风险分析"""
        sheet = self.wb[sheetName]
        riskReturn = {}
        # 读取第2行到59行的数据
        lines = range(2, 60)
        for i in lines:
            if sheet["A" + str(i)].value:
                riskReturn[sheet["A" + str(i)].value] = {
                    "所有交易": sheet["B" + str(i)].value,
                    "多头交易": sheet["C" + str(i)].value,
                    "空头交易": sheet["D" + str(i)].value
                }
        self.summaryDict.update({sheetName: riskReturn})
        return riskReturn

    def readTimeHoldingAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的时间仓位分析"""
        sheet = self.wb[sheetName]
        timeHolding = {}
        # 读取第2行到24行的数据
        lines = range(2, 24)
        for i in lines:
            if sheet["A" + str(i)].value:
                timeHolding[sheet["A" + str(i)].value] = {
                    "所有交易": sheet["B" + str(i)].value,
                    "多头交易": sheet["C" + str(i)].value,
                    "空头交易": sheet["D" + str(i)].value
                }
        self.summaryDict.update({sheetName: timeHolding})
        return timeHolding

    def getAllPictures(self):
        """
        """
        picHandle = jztProReportPicture(self.dirname, self.filename)
        picHandle.excel_pic_read()
        os.chdir(self.dirname + '\\' +
                 self.filename.lower().replace(".xlsx", "_pics"))
        print("renaming!!!")
        [
            os.rename(v + ".png", k + ".png")
            for k, v in self.labelImgdict.items()
        ]
        os.chdir(self.dirname)
        picHandle.clear()

    def readTradeDetails(self, sheetName=None):
        """读取金字塔专业回测报告中的交易明细"""
        details = pd.read_excel(open(self.fullpath, mode="rb"),
                                sheet_name=sheetName)
        self.summaryDict.update({sheetName: details})
        return details

    def readTradeOverallAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的交易总体分析"""
        sheet = self.wb[sheetName]
        overallAnalysis = {}
        # 读取第2行到31行的数据
        lines = range(2, 31)
        for i in lines:
            if sheet["A" + str(i)].value:
                overallAnalysis[sheet["A" + str(i)].value] = {
                    "所有交易": sheet["B" + str(i)].value,
                    "多头交易": sheet["C" + str(i)].value,
                    "空头交易": sheet["D" + str(i)].value
                }
        self.summaryDict.update({sheetName: overallAnalysis})
        return overallAnalysis

    def readContinuousPLAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的连续盈亏分析"""
        sheet = self.wb[sheetName]
        continuousPL = {}
        continuousP = {}
        continuousL = {}
        # 从第3行开始读取 ，出现空格结束查找
        lineNo = 0  # 用来标记连续盈利的结束位置
        lines = list(range(3, 100))  # 后面的100 没有意义
        for i in lines:
            if sheet["A" + str(i)].value:
                continuousP["连盈" + str(sheet["A" + str(i)].value) + "手"] = {
                    "出现次数": sheet["B" + str(i)].value,
                    "平均盈利": sheet["C" + str(i)].value,
                    "下一笔平均亏损": sheet["D" + str(i)].value
                }
            else:
                continuousPL["连续盈利手数"] = continuousP
                lineNo = i + 2
                break

        lines = list(range(lineNo, 100))  # 后面的100 没有意义
        for i in lines:
            if sheet["A" + str(i)].value:
                continuousL["连亏" + str(sheet["A" + str(i)].value) + "手"] = {
                    "出现次数": sheet["B" + str(i)].value,
                    "平均亏损": sheet["C" + str(i)].value,
                    "下一笔平均亏损": sheet["D" + str(i)].value
                }
            else:
                continuousPL["连续亏损手数"] = continuousL
                break
        self.summaryDict.update({sheetName: continuousPL})
        return continuousPL

    def readOutlierTrading(self, sheetName=None):
        """读取金字塔专业回测报告中的离群交易"""
        sheet = self.wb[sheetName]
        outliertrading = {}
        # 读取第2行到8行的数据
        lines = range(2, 9)
        for i in lines:
            if sheet["A" + str(i)].value:
                outliertrading[sheet["A" + str(i)].value] = {
                    "所有交易": sheet["B" + str(i)].value,
                    "多头交易": sheet["C" + str(i)].value,
                    "空头交易": sheet["D" + str(i)].value
                }
        self.summaryDict.update({sheetName: outliertrading})
        return outliertrading

    def readMaxFloatingPL(self, sheetName=None):
        """读取金字塔专业回测报告中的最大浮盈浮亏"""
        sheet = self.wb[sheetName]
        maxFloatingPL = {}
        # 读取第2行到9行的数据
        lines = range(2, 10)
        for i in lines:
            if sheet["A" + str(i)].value:
                maxFloatingPL[sheet["A" + str(i)].value] = {
                    "最大浮盈": sheet["B" + str(i)].value,
                    "最大浮亏": sheet["C" + str(i)].value
                }
        self.summaryDict.update({sheetName: maxFloatingPL})
        return maxFloatingPL

    def readDailyTradeAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的日交易分析"""
        dailyTrade = pd.read_excel(open(self.fullpath, mode="rb"),
                                   sheet_name=sheetName)
        self.summaryDict.update({sheetName: dailyTrade})
        return dailyTrade

    def readMonthlyPerformanceAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的月绩效分析"""
        monthlyPerformance = pd.read_excel(open(self.fullpath, mode="rb"),
                                           sheet_name=sheetName)
        self.summaryDict.update({sheetName: monthlyPerformance})
        return monthlyPerformance

    def readMonthlyAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的月度分析"""
        monthlyAnalysis = pd.read_excel(open(self.fullpath, mode="rb"),
                                        sheet_name=sheetName)
        self.summaryDict.update({sheetName: monthlyAnalysis})
        return monthlyAnalysis

    def readYearlyAnalysis(self, sheetName=None):
        """读取金字塔专业回测报告中的年度分析"""
        yearlyAnalysis = pd.read_excel(open(self.fullpath, mode="rb"),
                                       sheet_name=sheetName)
        self.summaryDict.update({sheetName: yearlyAnalysis})
        return yearlyAnalysis
